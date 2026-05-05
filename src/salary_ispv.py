"""ISPV XLSX loader — parses Czech ISPV M8r data files into SalaryData.

ISPV (Informační systém o průměrném výdělku) is the Czech Ministry of Labour
statistical dataset for occupational wage distributions, published annually.

This module reads the official ISPV M8r workbook format:
- Sheet "MZS-M8r" with multi-level merged headers in rows 0-7, data from row 8+.
- Column 0: label string with leading ISCO code (e.g. "1120 Nejvyšší...").
- Column 1: POCET in thousands. Cols 2-7: MEDIAN, D1, Q1, Q3, D9, PRUMER (CZK/měs).
- 4-digit rows are top-level subgroups; 5-digit rows (with leading space)
  are sub-categories — only 4-digit are indexed since the parser emits 4-digit codes.
- Maps decile columns to five seniority bands (junior … principal).
- Applies compound annual inflation correction per occupation family.
- Supports a 4→3→2→1 digit ISCO fallback chain before family-level fallback.
"""

from __future__ import annotations

import logging
import math
import re
from datetime import UTC, date, datetime
from pathlib import Path

from src.schemas import SalaryBand, SalaryData

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Publication date of the current ISPV annual dataset used for age calculation.
_ISPV_PUBLICATION_DATE = date(2026, 3, 1)

# Dataset version label embedded in every returned SalaryData.
_SOURCE_DATASET_VERSION = "ISPV-2025-annual"

# Minimum sample count to qualify for "high" confidence.
_HIGH_CONFIDENCE_MIN_COUNT = 30

# Real ISPV M8r layout: header band rows 0-7, data from row 8+.
# Sheet is named "MZS-M8r" (Mzdová sféra ČR — M8r table).
_M8R_SHEET_NAME = "MZS-M8r"
_M8R_DATA_START_ROW = 8  # 0-indexed; rows 0-7 are merged-cell title + headers

# Column positions in the M8r sheet (0-indexed).
_COL_LABEL = 0  # "1120 Nejvyšší představitelé..." — code + role name
_COL_POCET = 1  # tis. osob (thousands of employees)
_COL_MEDIAN = 2
_COL_D1 = 3
_COL_Q1 = 4
_COL_Q3 = 5
_COL_D9 = 6
_COL_PRUMER = 7

# Regex for extracting ISCO code + name from a label cell.
# 4-digit codes have no leading whitespace; 5-digit subgroups have one space.
_LABEL_RE = re.compile(r"^\s*(\d+)\s+(.+)$")

# ISCO 1-digit / 2-digit → occupation_family mapping.
# Lookup chain: 4-digit → 2-digit → 1-digit → "other".
_ISCO_FAMILY: dict[str, str] = {
    "1": "other",  # Managers
    "2": "it",  # Professionals — default; refined by 2-digit below
    "21": "it",  # Science and engineering
    "22": "healthcare",  # Health professionals
    "23": "education",  # Teaching professionals
    "24": "legal",  # Business and administration
    "26": "legal",  # Legal, social and cultural
    "3": "trades",  # Technicians and associate professionals
    "31": "it",  # Science and engineering technicians
    "32": "healthcare",  # Health associate professionals
    "5": "services",  # Service and sales workers
    "6": "trades",  # Skilled agricultural workers
    "7": "trades",  # Craft and related trades
    "8": "trades",  # Plant and machine operators
    "9": "services",  # Elementary occupations
}


# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------


class ISPVLookupError(Exception):
    """Base class for salary-estimation failures. Caught by UI/API.

    Subclasses encode the specific failure mode so callers can surface a
    targeted message (and HTTP status, in the API).
    """


class ISPVDataMissingError(ISPVLookupError):
    """Raised when the ISPV dataset is not loaded or empty."""


class NonCZLocationError(ISPVLookupError):
    """Raised when the candidate's location is outside the CZ market."""


class MissingISCOError(ISPVLookupError):
    """Raised when the parsed CV contains no ISCO code on any experience."""


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _isco_family(isco_code: str) -> str:
    """Resolve occupation family from ISCO code using longest-prefix lookup.

    Args:
        isco_code: 1-4 digit ISCO code string.

    Returns:
        Occupation family string; "other" if no match found.
    """
    # Try 2-digit prefix, then 1-digit prefix.
    for length in (2, 1):
        prefix = isco_code[:length]
        if prefix in _ISCO_FAMILY:
            return _ISCO_FAMILY[prefix]
    return "other"


def _round_czk(value: float) -> int:
    """Round a CZK float value to whole CZK.

    Args:
        value: Float CZK amount.

    Returns:
        Nearest integer CZK value.
    """
    return round(value)


def _build_bands(
    d1: float,
    q1: float,
    median: float,
    q3: float,
    d9: float,
    inflation_factor: float,
) -> tuple[SalaryBand, SalaryBand, SalaryBand, SalaryBand, SalaryBand]:
    """Compute five seniority bands from ISPV decile values with inflation applied.

    Mapping:
        junior    D1  .. Q1         (raw percentile range)
        medior    Q1  .. Q3 mid     (Q1 to midpoint of Q3)
        senior    Q3 mid .. D9
        lead      D9 × [1.10, 1.20, 1.35]  (extrapolated above D9)
        principal D9 × [1.30, 1.50, 1.75]

    Args:
        d1:               10th percentile wage (CZK/month).
        q1:               25th percentile wage.
        median:           50th percentile wage.
        q3:               75th percentile wage.
        d9:               90th percentile wage.
        inflation_factor: Compound multiplier to apply to all values.

    Returns:
        Tuple of (junior, medior, senior, lead, principal) SalaryBand objects.
    """

    def _apply(value: float) -> int:
        return _round_czk(value * inflation_factor)

    junior = SalaryBand(
        low=_apply(d1),
        mid=_apply((d1 + q1) / 2),
        high=_apply(q1),
    )
    medior = SalaryBand(
        low=_apply(q1),
        mid=_apply(median),
        high=_apply((median + q3) / 2),
    )
    senior = SalaryBand(
        low=_apply((median + q3) / 2),
        mid=_apply(q3),
        high=_apply(d9),
    )
    lead = SalaryBand(
        low=_apply(d9 * 1.10),
        mid=_apply(d9 * 1.20),
        high=_apply(d9 * 1.35),
    )
    principal = SalaryBand(
        low=_apply(d9 * 1.30),
        mid=_apply(d9 * 1.50),
        high=_apply(d9 * 1.75),
    )
    return junior, medior, senior, lead, principal


def _dataset_age_months(reference_date: date | None = None) -> int:
    """Compute integer months elapsed since _ISPV_PUBLICATION_DATE.

    Args:
        reference_date: Override today's date (for testing). Defaults to today.

    Returns:
        Non-negative integer count of elapsed months.
    """
    today = reference_date or datetime.now(tz=UTC).date()
    delta_days = (today - _ISPV_PUBLICATION_DATE).days
    return max(0, math.floor(delta_days / 30.4375))  # ~365.25/12 days per month


def _compound_factor(annual_rate: float, age_months: int) -> float:
    """Compute compound inflation factor for given months.

    Formula: annual_rate ** (age_months / 12)

    Args:
        annual_rate: Annual inflation multiplier (e.g. 1.05 for 5 % p.a.).
        age_months:  Number of months to compound over.

    Returns:
        Compound multiplier >= 1.0.
    """
    if age_months <= 0:
        return 1.0
    return annual_rate ** (age_months / 12)


# ---------------------------------------------------------------------------
# Row record (internal)
# ---------------------------------------------------------------------------


class _ISPVRow:
    """Parsed row from the ISPV XLSX sheet."""

    __slots__ = ("isco_code", "d1", "q1", "median", "prumer", "q3", "d9", "pocet", "name")

    def __init__(
        self,
        isco_code: str,
        d1: float,
        q1: float,
        median: float,
        prumer: float,
        q3: float,
        d9: float,
        pocet: int,
        name: str | None,
    ) -> None:
        self.isco_code = isco_code
        self.d1 = d1
        self.q1 = q1
        self.median = median
        self.prumer = prumer
        self.q3 = q3
        self.d9 = d9
        self.pocet = pocet
        self.name = name


# ---------------------------------------------------------------------------
# ISPVLoader
# ---------------------------------------------------------------------------


class ISPVLoader:
    """Loads an ISPV M8r XLSX file and provides ISCO-code salary lookup.

    Usage::

        loader = ISPVLoader(xlsx_path=Path("ispv_2025.xlsx"), inflation_factors={...})
        loader.load()
        salary_data = loader.lookup("2512")

    The lookup follows a fallback chain:
        4-digit exact → 3-digit prefix → 2-digit prefix → 1-digit prefix → family fallback.

    If the KRAJ column is absent, all rows are treated as national aggregate data.
    """

    def __init__(
        self,
        xlsx_path: Path,
        inflation_factors: dict[str, float],
    ) -> None:
        """Initialise loader with file path and inflation configuration.

        Args:
            xlsx_path:         Path to the ISPV XLSX file to parse.
            inflation_factors: Mapping of occupation_family → annual CPI multiplier.
                               Use key "other" as the default fallback rate.
        """
        self._xlsx_path = xlsx_path
        self._inflation_factors = inflation_factors
        # Mapping from 4-digit ISCO code to parsed row — populated by load().
        self._data: dict[str, _ISPVRow] = {}
        self._loaded = False

    # ------------------------------------------------------------------
    # Public interface
    # ------------------------------------------------------------------

    def is_loaded(self) -> bool:
        """Return True if load() has been called successfully.

        Returns:
            Boolean indicating whether the dataset is ready for lookup.
        """
        return self._loaded

    def known_codes(self) -> frozenset[str]:
        """Return frozenset of all 4-digit ISCO codes currently indexed."""
        return frozenset(self._data.keys())

    def best_code_for_family(self, family: str) -> str | None:
        """Return the 4-digit ISCO code with highest POCET in the given family, or None."""
        best: _ISPVRow | None = None
        for row in self._data.values():
            if _isco_family(row.isco_code) == family and (best is None or row.pocet > best.pocet):
                best = row
        return best.isco_code if best else None

    def load(self) -> None:
        """Parse the XLSX file and populate the internal index.

        Reads the "MZS-M8r" sheet (or the first sheet as fallback), skips the
        merged-header band (rows 0-7), and parses data rows positionally.
        Indexes rows whose label starts with a 4-digit ISCO code; 5-digit
        sub-categories (with leading whitespace in label) are skipped because
        the parser emits 4-digit codes only.

        Raises:
            FileNotFoundError: If xlsx_path does not exist.
            ValueError:        If no data rows can be parsed (file format mismatch).
        """
        import openpyxl  # local import — optional dependency

        logger.info("Loading ISPV dataset from %s", self._xlsx_path)
        if not self._xlsx_path.exists():
            raise FileNotFoundError(f"ISPV XLSX not found: {self._xlsx_path}")

        wb = openpyxl.load_workbook(self._xlsx_path, read_only=True, data_only=True)
        ws = wb[_M8R_SHEET_NAME] if _M8R_SHEET_NAME in wb.sheetnames else wb.worksheets[0]

        parsed: dict[str, _ISPVRow] = {}
        skipped = 0

        for row_idx, raw_row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx < _M8R_DATA_START_ROW:
                continue

            label = raw_row[_COL_LABEL]
            if not isinstance(label, str):
                skipped += 1
                continue

            # Skip 5-digit sub-categories (leading whitespace in label).
            # Only top-level 4-digit groups are indexed — parser emits 4-digit codes.
            if label.startswith(" "):
                skipped += 1
                continue

            match = _LABEL_RE.match(label)
            if match is None:
                skipped += 1
                continue

            isco_code = match.group(1)
            role_name = match.group(2).strip()

            if len(isco_code) != 4 or isco_code[0] == "0":
                skipped += 1
                continue

            # openpyxl returns _CellGetValue for read-only cells; coerce via str()
            # before float() — strict typing-friendly and accepts ints, floats, str numerics.
            try:
                pocet_thousands = float(str(raw_row[_COL_POCET]))
                median = float(str(raw_row[_COL_MEDIAN]))
                d1 = float(str(raw_row[_COL_D1]))
                q1 = float(str(raw_row[_COL_Q1]))
                q3 = float(str(raw_row[_COL_Q3]))
                d9 = float(str(raw_row[_COL_D9]))
                prumer = float(str(raw_row[_COL_PRUMER]))
            except (TypeError, ValueError, IndexError):
                logger.debug("Skipping non-numeric row for ISCO %s", isco_code)
                skipped += 1
                continue

            parsed[isco_code] = _ISPVRow(
                isco_code=isco_code,
                d1=d1,
                q1=q1,
                median=median,
                prumer=prumer,
                q3=q3,
                d9=d9,
                pocet=int(round(pocet_thousands * 1000)),  # tis. osob → count
                name=role_name,
            )

        wb.close()
        if not parsed:
            raise ValueError(
                f"ISPV XLSX at {self._xlsx_path} produced 0 indexed rows — "
                "format may have changed (expected sheet 'MZS-M8r' with data from row 9)."
            )

        self._data = parsed
        self._loaded = True
        logger.info(
            "ISPV load complete: %d 4-digit entries indexed, %d rows skipped",
            len(parsed),
            skipped,
        )

    def lookup(self, isco_code: str) -> SalaryData:
        """Return SalaryData for the given ISCO code, using fallback chain if needed.

        Lookup order:
            1. 4-digit exact match
            2. 3-digit prefix match (first matching entry)
            3. 2-digit prefix match
            4. 1-digit prefix match
            5. Occupation family fallback (any entry sharing the same family)

        Args:
            isco_code: CZ-ISCO-08 code string, 1-4 digits.

        Returns:
            SalaryData populated from matched ISPV row with inflation applied.

        Raises:
            RuntimeError:      If load() has not been called yet.
            ValueError:        If isco_code format is invalid.
            ISPVLookupError:   If no matching row can be found at any fallback level.
        """
        if not self._loaded:
            raise RuntimeError("ISPVLoader.load() must be called before lookup()")

        # Validate input format: 1-4 digits, first digit 1-9.
        if not isco_code.isdigit() or not (1 <= len(isco_code) <= 4) or isco_code[0] == "0":
            raise ValueError(
                f"Invalid ISCO code format: {isco_code!r}. Must be 1-4 digits with first digit 1-9."
            )

        row, match_level = self._resolve(isco_code)
        if row is None:
            raise ISPVLookupError(
                f"No ISPV data found for ISCO code {isco_code!r} at any fallback level"
            )

        return self._build_salary_data(isco_code, row, match_level)

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    def _resolve(self, isco_code: str) -> tuple[_ISPVRow | None, str]:
        """Try to find a matching row using the full fallback chain.

        Args:
            isco_code: Input ISCO code (already validated).

        Returns:
            Tuple of (matched row or None, match_level string).
        """
        # 1. 4-digit exact
        if isco_code in self._data:
            return self._data[isco_code], "4digit"

        # 2. 3-digit prefix
        prefix3 = isco_code[:3]
        match3 = self._find_prefix(prefix3)
        if match3 is not None:
            return match3, "3digit"

        # 3. 2-digit prefix
        prefix2 = isco_code[:2]
        match2 = self._find_prefix(prefix2)
        if match2 is not None:
            return match2, "2digit"

        # 4. 1-digit prefix
        prefix1 = isco_code[:1]
        match1 = self._find_prefix(prefix1)
        if match1 is not None:
            return match1, "1digit"

        # 5. Family fallback — entry with highest POCET in same occupation family.
        target_family = _isco_family(isco_code)
        best: _ISPVRow | None = None
        for candidate in self._data.values():
            if _isco_family(candidate.isco_code) == target_family and (
                best is None or candidate.pocet > best.pocet
            ):
                best = candidate
        if best is not None:
            logger.warning(
                "ISCO %s: family fallback to %s (family=%s, pocet=%d)",
                isco_code,
                best.isco_code,
                target_family,
                best.pocet,
            )
            return best, "family_fallback"

        return None, "family_fallback"

    def _find_prefix(self, prefix: str) -> _ISPVRow | None:
        """Return the highest-POCET row whose ISCO code starts with prefix, or None.

        Picking by sample size makes the fallback deterministic and selects the
        most representative occupation in the prefix group rather than whichever
        row happens to come first in the source XLSX.
        """
        best: _ISPVRow | None = None
        for code, row in self._data.items():
            if code.startswith(prefix) and (best is None or row.pocet > best.pocet):
                best = row
        return best

    def _build_salary_data(
        self,
        input_code: str,
        row: _ISPVRow,
        match_level: str,
    ) -> SalaryData:
        """Construct a SalaryData from a resolved ISPV row.

        Args:
            input_code:  The original ISCO code supplied by the caller.
            row:         The matched ISPV row.
            match_level: One of "4digit", "3digit", "2digit", "1digit", "family_fallback".

        Returns:
            Fully populated SalaryData with inflation applied.
        """
        family = _isco_family(row.isco_code)
        annual_rate = self._inflation_factors.get(family, self._inflation_factors.get("other", 1.0))
        age_months = _dataset_age_months()
        factor = _compound_factor(annual_rate, age_months)

        junior, medior, senior, lead, principal = _build_bands(
            d1=row.d1,
            q1=row.q1,
            median=row.median,
            q3=row.q3,
            d9=row.d9,
            inflation_factor=factor,
        )

        # Determine confidence level.
        warnings: list[str] = []
        if row.pocet < _HIGH_CONFIDENCE_MIN_COUNT:
            warnings.append(
                f"Low sample count: POCET={row.pocet} < {_HIGH_CONFIDENCE_MIN_COUNT}; "
                "band estimates are less reliable."
            )

        if match_level == "4digit" and row.pocet >= _HIGH_CONFIDENCE_MIN_COUNT:
            confidence: str = "high"
        elif match_level in ("3digit", "2digit") or row.pocet < _HIGH_CONFIDENCE_MIN_COUNT:
            confidence = "medium"
        else:
            confidence = "low"

        if match_level in ("1digit", "family_fallback"):
            confidence = "low"
            warnings.append(
                f"Imprecise ISCO match: {match_level} fallback used for input code {input_code!r}."
            )

        role_name = row.name or f"ISCO {row.isco_code}"
        role_slug = f"isco_{row.isco_code}"

        # The matched 4-digit code is stored; if we did a prefix fallback the stored
        # code comes from the matched row (which is already a 4-digit entry).
        matched_code = row.isco_code

        return SalaryData(
            role_name=role_name,
            role_slug=role_slug,
            researched_at=datetime.now(tz=UTC),
            ttl_days=90,
            junior=junior,
            medior=medior,
            senior=senior,
            lead=lead,
            principal=principal,
            isco_code=matched_code,
            isco_match_level=match_level,  # type: ignore[arg-type]
            dataset_age_months=age_months,
            inflation_factor_applied=factor,
            source_dataset_version=_SOURCE_DATASET_VERSION,
            confidence=confidence,  # type: ignore[arg-type]
            warnings=warnings,
        )
