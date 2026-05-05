"""Unit tests for src/salary_ispv.py — ISPVLoader with synthetic XLSX fixture.

All tests use the ispv_sample_xlsx fixture from conftest.py and never touch
live ISPV files or external APIs.

Fixture layout (MZS-M8r real format):
    Rows 0-7:  header band (skipped by loader)
    Row 8+:    data rows — 5 valid 4-digit entries + 1 skipped 5-digit subgroup

    2512 Vývojáři softwaru         POCET=0.450 → 450  (high confidence)
    2211 Lékaři specialisté        POCET=0.200 → 200  (high confidence)
    2612 Právníci                  POCET=0.025 →  25  (low sample < 30)
    3112 Stavební technici         POCET=0.150 → 150  (high confidence)
    9112 Pomocníci úklid           POCET=0.300 → 300  (high confidence)
    " 11201 Subgroup test"         SKIPPED — 5-digit subgroup with leading space
"""

from __future__ import annotations

import math
from pathlib import Path
from unittest.mock import patch

import pytest

from src.salary_ispv import (
    ISPVLoader,
    ISPVLookupError,
    _compound_factor,  # pyright: ignore[reportPrivateUsage]
    _dataset_age_months,  # pyright: ignore[reportPrivateUsage]
    _isco_family,  # pyright: ignore[reportPrivateUsage]
)
from src.schemas import SalaryData

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_DEFAULT_INFLATION: dict[str, float] = {
    "it": 1.05,
    "healthcare": 1.04,
    "legal": 1.03,
    "education": 1.02,
    "trades": 1.03,
    "services": 1.02,
    "other": 1.02,
}


def _make_loader(xlsx_path: Path, inflation: dict[str, float] | None = None) -> ISPVLoader:
    """Return a loaded ISPVLoader using the given fixture path."""
    loader = ISPVLoader(
        xlsx_path=xlsx_path,
        inflation_factors=inflation or _DEFAULT_INFLATION,
    )
    loader.load()
    return loader


def _m8r_workbook(
    tmp_path: Path, data_rows: list[list[object]], sheet_name: str = "MZS-M8r"
) -> Path:
    """Create a minimal M8r XLSX with 8 header rows + given data rows.

    Args:
        tmp_path:   pytest tmp_path fixture value.
        data_rows:  Each item is a list [label, pocet, median, d1, q1, q3, d9, prumer].
        sheet_name: Override worksheet title (default: MZS-M8r).

    Returns:
        Path to the saved XLSX file.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet_name
    for _ in range(8):
        ws.append(["HEADER", None, None, None, None, None, None, None])
    for row in data_rows:
        ws.append(row)
    path = tmp_path / "test.xlsx"
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# load() — parsing: correct row count
# ---------------------------------------------------------------------------


def test_load_parses_5_rows(ispv_sample_xlsx: Path) -> None:
    """Five 4-digit entries must be indexed; the 5-digit subgroup must be excluded."""
    loader = _make_loader(ispv_sample_xlsx)
    assert len(loader._data) == 5  # noqa: SLF001  # pyright: ignore[reportPrivateUsage]


def test_load_sets_is_loaded(ispv_sample_xlsx: Path) -> None:
    """is_loaded() must return False before load() and True after."""
    loader = ISPVLoader(xlsx_path=ispv_sample_xlsx, inflation_factors=_DEFAULT_INFLATION)
    assert loader.is_loaded() is False
    loader.load()
    assert loader.is_loaded() is True


# ---------------------------------------------------------------------------
# load() — 5-digit subgroup skip
# ---------------------------------------------------------------------------


def test_load_skips_5digit_subgroups(ispv_sample_xlsx: Path) -> None:
    """Label starting with a space (' 11201 ...') must not appear in _data."""
    loader = _make_loader(ispv_sample_xlsx)
    assert "11201" not in loader._data  # noqa: SLF001  # pyright: ignore[reportPrivateUsage]
    # None of the keys in _data are 5-digit codes.
    for code in loader._data:  # noqa: SLF001  # pyright: ignore[reportPrivateUsage]
        assert len(code) == 4, f"Expected 4-digit code, got {code!r}"


def test_load_skips_label_without_leading_digits(tmp_path: Path) -> None:
    """Rows whose label does not start with digits must be silently skipped."""
    path = _m8r_workbook(
        tmp_path,
        [
            ["Popis skupiny", 0.1, 50000, 30000, 40000, 60000, 75000, 52000],  # no ISCO prefix
            ["2512 Vývojáři softwaru", 0.45, 78000, 45000, 60000, 100000, 130000, 82000],
        ],
    )
    loader = _make_loader(path)
    assert len(loader._data) == 1  # noqa: SLF001  # pyright: ignore[reportPrivateUsage]
    assert "2512" in loader._data  # noqa: SLF001  # pyright: ignore[reportPrivateUsage]


def test_load_skips_rows_without_isco_label(tmp_path: Path) -> None:
    """Rows where the label cell is None (not a string) must be silently skipped."""
    path = _m8r_workbook(
        tmp_path,
        [
            [None, 0.1, 50000, 30000, 40000, 60000, 75000, 52000],  # None label
            ["2512 Vývojáři softwaru", 0.45, 78000, 45000, 60000, 100000, 130000, 82000],
        ],
    )
    loader = _make_loader(path)
    assert len(loader._data) == 1  # noqa: SLF001  # pyright: ignore[reportPrivateUsage]
    assert "2512" in loader._data  # noqa: SLF001  # pyright: ignore[reportPrivateUsage]


# ---------------------------------------------------------------------------
# load() — sheet name selection
# ---------------------------------------------------------------------------


def test_load_uses_mzs_m8r_sheet_name(tmp_path: Path) -> None:
    """When the workbook has multiple sheets, 'MZS-M8r' must be preferred over others."""
    from openpyxl import Workbook

    wb = Workbook()
    # First sheet: dummy content that would yield no valid rows.
    ws_dummy = wb.active
    assert ws_dummy is not None
    ws_dummy.title = "Titulní list"
    for _ in range(8):
        ws_dummy.append(["HEADER", None, None, None, None, None, None, None])
    ws_dummy.append(["Not a real row", 0.0, 0, 0, 0, 0, 0, 0])

    # Second sheet: real M8r data.
    ws_real = wb.create_sheet(title="MZS-M8r")
    for _ in range(8):
        ws_real.append(["HEADER", None, None, None, None, None, None, None])
    ws_real.append(["2512 Vývojáři softwaru", 0.45, 78000, 45000, 60000, 100000, 130000, 82000])

    path = tmp_path / "multi_sheet.xlsx"
    wb.save(path)

    loader = _make_loader(path)
    assert "2512" in loader._data  # noqa: SLF001  # pyright: ignore[reportPrivateUsage]


# ---------------------------------------------------------------------------
# load() — POCET conversion from thousands
# ---------------------------------------------------------------------------


def test_load_pocet_converted_from_thousands(ispv_sample_xlsx: Path) -> None:
    """POCET=0.450 in the XLSX must be stored internally as 450 (persons, not thousands)."""
    loader = _make_loader(ispv_sample_xlsx)
    row = loader._data["2512"]  # noqa: SLF001  # pyright: ignore[reportPrivateUsage]
    assert row.pocet == 450


# ---------------------------------------------------------------------------
# load() — numeric rows skipped silently
# ---------------------------------------------------------------------------


def test_load_skips_non_numeric_rows(tmp_path: Path) -> None:
    """Rows with non-numeric decile values must be silently skipped."""
    path = _m8r_workbook(
        tmp_path,
        [
            ["2512 Vývojáři", "n/a", "n/a", "n/a", "n/a", "n/a", "n/a", "n/a"],  # all non-numeric
            ["3112 Technici", 0.15, 55000, 35000, 44000, 68000, 85000, 58000],  # valid
        ],
    )
    loader = _make_loader(path)
    assert len(loader._data) == 1  # noqa: SLF001  # pyright: ignore[reportPrivateUsage]
    assert "3112" in loader._data  # noqa: SLF001  # pyright: ignore[reportPrivateUsage]


# ---------------------------------------------------------------------------
# load() — error paths
# ---------------------------------------------------------------------------


def test_load_raises_on_empty_xlsx(tmp_path: Path) -> None:
    """Completely empty XLSX (no data rows after header band) must raise ValueError."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "MZS-M8r"
    # Write only the 8 header rows — no data rows.
    for _ in range(8):
        ws.append(["HEADER", None, None, None, None, None, None, None])
    path = tmp_path / "empty_sheet.xlsx"
    wb.save(path)

    loader = ISPVLoader(xlsx_path=path, inflation_factors=_DEFAULT_INFLATION)
    with pytest.raises(ValueError):
        loader.load()


def test_load_raises_on_empty_data_rows(tmp_path: Path) -> None:
    """XLSX with 8 header rows and no valid data rows must raise ValueError."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "MZS-M8r"
    for _ in range(8):
        ws.append(["HEADER", None, None, None, None, None, None, None])
    # One data row but label is not an ISCO entry — will be skipped.
    ws.append(["Not an ISCO label", 0.1, 50000, 30000, 40000, 60000, 75000, 52000])
    path = tmp_path / "no_valid_data.xlsx"
    wb.save(path)

    loader = ISPVLoader(xlsx_path=path, inflation_factors=_DEFAULT_INFLATION)
    with pytest.raises(ValueError):
        loader.load()


# ---------------------------------------------------------------------------
# lookup() — 4-digit exact match
# ---------------------------------------------------------------------------


def test_lookup_4digit_exact_match(ispv_sample_xlsx: Path) -> None:
    """ISCO 2512 must match exactly at '4digit' level."""
    loader = _make_loader(ispv_sample_xlsx)
    result: SalaryData = loader.lookup("2512")

    assert isinstance(result, SalaryData)
    assert result.isco_code == "2512"
    assert result.isco_match_level == "4digit"
    assert result.role_slug == "isco_2512"
    # All five seniority bands must be populated.
    assert result.junior is not None
    assert result.medior is not None
    assert result.senior is not None
    assert result.lead is not None
    assert result.principal is not None


def test_lookup_role_name_parsed_from_label(ispv_sample_xlsx: Path) -> None:
    """role_name must be extracted from the label string (not a separate column)."""
    loader = _make_loader(ispv_sample_xlsx)
    result = loader.lookup("2512")
    assert result.role_name == "Vývojáři softwaru"


def test_lookup_confidence_high_for_large_sample(ispv_sample_xlsx: Path) -> None:
    """4-digit match with POCET=450 (>= 30) must yield confidence='high'."""
    loader = _make_loader(ispv_sample_xlsx)
    result = loader.lookup("2512")
    assert result.confidence == "high"


# ---------------------------------------------------------------------------
# lookup() — 3-digit prefix fallback
# ---------------------------------------------------------------------------


def test_lookup_3digit_fallback(ispv_sample_xlsx: Path) -> None:
    """Code '2519' not in dataset should fall back to the 3-digit '251X' match."""
    loader = _make_loader(ispv_sample_xlsx)
    # 2519 doesn't exist; 2512 starts with '251', so it should match.
    result = loader.lookup("2519")

    assert result.isco_match_level == "3digit"
    # The matched row comes from the "251" prefix — code 2512.
    assert result.isco_code == "2512"


# ---------------------------------------------------------------------------
# lookup() — family fallback
# ---------------------------------------------------------------------------


def test_lookup_family_fallback(ispv_sample_xlsx: Path) -> None:
    """Code '9999' (unknown 4-digit) must eventually fall back via 1-digit or family."""
    loader = _make_loader(ispv_sample_xlsx)
    # 9999: 3-digit "999" absent; 2-digit "99" absent; 1-digit "9" matches "9112" entry.
    result = loader.lookup("9999")

    assert result.isco_match_level in ("1digit", "family_fallback")
    assert result.confidence == "low"


def test_lookup_family_fallback_adds_warning(ispv_sample_xlsx: Path) -> None:
    """family_fallback (or 1-digit) match must produce at least one warning string.

    Code '5999' has no 4/3/2/1-digit prefix match in the fixture
    (no 5xxx entries). Its occupation family is 'services', shared with '9112',
    triggering at least a 1-digit or family-level fallback.
    """
    loader = _make_loader(ispv_sample_xlsx)
    result = loader.lookup("5999")
    assert result.confidence == "low"
    assert len(result.warnings) >= 1


# ---------------------------------------------------------------------------
# Decile → band mapping
# ---------------------------------------------------------------------------


def test_decile_mapping_junior_band(ispv_sample_xlsx: Path) -> None:
    """junior.low must equal D1 and junior.high must equal Q1 (inflation=1.0)."""
    # Use factor=1.0 so inflation doesn't obscure the mapping.
    loader = _make_loader(ispv_sample_xlsx, inflation={"it": 1.0, "other": 1.0})

    # Freeze dataset_age_months to 0 so compound factor = 1.0.
    with patch("src.salary_ispv._dataset_age_months", return_value=0):
        result = loader.lookup("2512")

    # Fixture 2512: D1=45000, Q1=60000
    assert result.junior is not None
    assert result.junior.low == 45000
    assert result.junior.high == 60000
    assert result.junior.mid == round((45000 + 60000) / 2)


def test_decile_mapping_principal_extrapolated(ispv_sample_xlsx: Path) -> None:
    """principal band must use D9 × [1.30, 1.50, 1.75] multipliers."""
    loader = _make_loader(ispv_sample_xlsx, inflation={"it": 1.0, "other": 1.0})

    with patch("src.salary_ispv._dataset_age_months", return_value=0):
        result = loader.lookup("2512")

    # Fixture 2512: D9=130000
    d9 = 130000
    assert result.principal is not None
    assert result.principal.low == round(d9 * 1.30)
    assert result.principal.mid == round(d9 * 1.50)
    assert result.principal.high == round(d9 * 1.75)


def test_decile_mapping_lead_extrapolated(ispv_sample_xlsx: Path) -> None:
    """lead band must use D9 × [1.10, 1.20, 1.35] multipliers."""
    loader = _make_loader(ispv_sample_xlsx, inflation={"it": 1.0, "other": 1.0})

    with patch("src.salary_ispv._dataset_age_months", return_value=0):
        result = loader.lookup("2512")

    d9 = 130000
    assert result.lead is not None
    assert result.lead.low == round(d9 * 1.10)
    assert result.lead.mid == round(d9 * 1.20)
    assert result.lead.high == round(d9 * 1.35)


# ---------------------------------------------------------------------------
# Inflation compounding
# ---------------------------------------------------------------------------


def test_inflation_factor_compounded(ispv_sample_xlsx: Path) -> None:
    """6-month inflation at 10% annual should apply factor = 1.10 ** 0.5."""
    inflation = {"it": 1.10, "other": 1.10}
    loader = _make_loader(ispv_sample_xlsx, inflation=inflation)

    with patch("src.salary_ispv._dataset_age_months", return_value=6):
        result = loader.lookup("2512")

    expected_factor = 1.10**0.5  # ~1.0488
    # junior.low is D1 * factor = 45000 * expected_factor
    expected_low = round(45000 * expected_factor)
    assert result.junior is not None
    assert result.junior.low == expected_low


def test_compound_factor_zero_months() -> None:
    """Zero months must return factor 1.0 (no inflation applied)."""
    assert _compound_factor(1.10, 0) == 1.0


def test_compound_factor_12_months() -> None:
    """Exactly 12 months at rate=1.05 must return 1.05 (one full year)."""
    result = _compound_factor(1.05, 12)
    assert math.isclose(result, 1.05, rel_tol=1e-9)


def test_dataset_age_months_nonnegative() -> None:
    """dataset_age_months must never return a negative value."""
    age = _dataset_age_months()
    assert age >= 0


# ---------------------------------------------------------------------------
# Low sample count warning
# ---------------------------------------------------------------------------


def test_low_sample_warning(ispv_sample_xlsx: Path) -> None:
    """POCET < 30 (fixture: 2612 has 25 persons) must add a warning and confidence != 'high'."""
    loader = _make_loader(ispv_sample_xlsx)
    # 2612 has POCET=0.025 → 25 persons in the fixture — below the 30-person threshold.
    result = loader.lookup("2612")

    assert any("25" in w for w in result.warnings), (
        f"Expected low-sample warning mentioning count 25, got: {result.warnings}"
    )
    assert result.confidence in ("medium", "low")


# ---------------------------------------------------------------------------
# Error paths
# ---------------------------------------------------------------------------


def test_lookup_raises_when_not_loaded(ispv_sample_xlsx: Path) -> None:
    """Calling lookup() before load() must raise RuntimeError."""
    loader = ISPVLoader(xlsx_path=ispv_sample_xlsx, inflation_factors=_DEFAULT_INFLATION)
    with pytest.raises(RuntimeError, match="load\\(\\)"):
        loader.lookup("2512")


def test_invalid_isco_format_raises_non_digit(ispv_sample_xlsx: Path) -> None:
    """Non-digit ISCO code must raise ValueError."""
    loader = _make_loader(ispv_sample_xlsx)
    with pytest.raises(ValueError, match="Invalid ISCO"):
        loader.lookup("abcd")


def test_invalid_isco_format_raises_leading_zero(ispv_sample_xlsx: Path) -> None:
    """ISCO code starting with '0' must raise ValueError."""
    loader = _make_loader(ispv_sample_xlsx)
    with pytest.raises(ValueError, match="Invalid ISCO"):
        loader.lookup("0123")


def test_invalid_isco_format_raises_too_long(ispv_sample_xlsx: Path) -> None:
    """5-digit ISCO code must raise ValueError (max 4 digits)."""
    loader = _make_loader(ispv_sample_xlsx)
    with pytest.raises(ValueError, match="Invalid ISCO"):
        loader.lookup("01234")


def test_lookup_missing_file_raises(tmp_path: Path) -> None:
    """load() with non-existent path must raise FileNotFoundError."""
    loader = ISPVLoader(
        xlsx_path=tmp_path / "nonexistent.xlsx",
        inflation_factors=_DEFAULT_INFLATION,
    )
    with pytest.raises(FileNotFoundError):
        loader.load()


def test_lookup_no_match_raises_ispv_error(tmp_path: Path) -> None:
    """When no ISCO family match exists, ISPVLookupError must be raised.

    Only ISCO 2512 (family 'it') is in the dataset.
    ISCO 1111 has family 'other' — no 'other' entry exists, so the full
    fallback chain (4→3→2→1→family) fails and ISPVLookupError is raised.
    """
    path = _m8r_workbook(
        tmp_path,
        [["2512 Vývojáři softwaru", 0.45, 78000, 45000, 60000, 100000, 130000, 82000]],
    )
    loader = _make_loader(path)
    # 1111: family='other' — nothing in dataset shares that family.
    with pytest.raises(ISPVLookupError):
        loader.lookup("1111")


# ---------------------------------------------------------------------------
# Source dataset version
# ---------------------------------------------------------------------------


def test_source_dataset_version(ispv_sample_xlsx: Path) -> None:
    """source_dataset_version must match the hardcoded constant."""
    loader = _make_loader(ispv_sample_xlsx)
    result = loader.lookup("2512")
    assert result.source_dataset_version == "ISPV-2025-annual"


# ---------------------------------------------------------------------------
# Inflation factor stored in result
# ---------------------------------------------------------------------------


def test_inflation_factor_applied_stored(ispv_sample_xlsx: Path) -> None:
    """inflation_factor_applied must equal compound factor for the resolved family."""
    loader = _make_loader(ispv_sample_xlsx, inflation={"it": 1.05, "other": 1.02})

    with patch("src.salary_ispv._dataset_age_months", return_value=6):
        result = loader.lookup("2512")

    expected = 1.05 ** (6 / 12)
    assert math.isclose(result.inflation_factor_applied, expected, rel_tol=1e-9)


def test_inflation_factor_applied_ge_1(ispv_sample_xlsx: Path) -> None:
    """inflation_factor_applied must always be >= 1.0 (schema constraint)."""
    loader = _make_loader(ispv_sample_xlsx)
    result = loader.lookup("2512")
    assert result.inflation_factor_applied >= 1.0


# ---------------------------------------------------------------------------
# _isco_family() helper
# ---------------------------------------------------------------------------


def test_isco_family_fallback_returns_other_for_unknown_prefix() -> None:
    """_isco_family() must return 'other' for ISCO codes with no known 2/1-digit prefix."""
    # ISCO major group "4" (Clerical support workers) is not in _ISCO_FAMILY.
    result = _isco_family("4111")
    assert result == "other"


def test_isco_family_2digit_it_override() -> None:
    """ISCO 21xx (Science and engineering professionals) must resolve to 'it'."""
    assert _isco_family("2151") == "it"


def test_isco_family_healthcare() -> None:
    """ISCO 22xx (Health professionals) must resolve to 'healthcare'."""
    assert _isco_family("2211") == "healthcare"


# ---------------------------------------------------------------------------
# known_codes() — Bug 1 whitelist
# ---------------------------------------------------------------------------


def test_known_codes_returns_loaded_set(ispv_sample_xlsx: Path) -> None:
    """known_codes() must return a frozenset containing every indexed 4-digit code."""
    loader = _make_loader(ispv_sample_xlsx)
    codes = loader.known_codes()

    assert isinstance(codes, frozenset)
    assert "2512" in codes
    assert "2211" in codes
    assert "2612" in codes
    assert "3112" in codes
    assert "9112" in codes
    # 5-digit subgroup must NOT be present
    assert "11201" not in codes


# ---------------------------------------------------------------------------
# best_code_for_family() — Bug 1 whitelist
# ---------------------------------------------------------------------------


def test_best_code_for_family_returns_highest_pocet(tmp_path: Path) -> None:
    """best_code_for_family() must return the code with the highest POCET in the given family."""
    # Two IT codes with different POCET: 2512 (POCET=0.450 → 450) and 2514 (POCET=0.100 → 100).
    # Expect 2512 to win.
    path = _m8r_workbook(
        tmp_path,
        [
            ["2512 Vývojáři softwaru", 0.450, 78000, 45000, 60000, 100000, 130000, 82000],
            ["2514 Databázoví správci", 0.100, 65000, 40000, 52000, 85000, 110000, 70000],
        ],
    )
    loader = _make_loader(path)
    result = loader.best_code_for_family("it")

    assert result == "2512"


def test_best_code_for_family_returns_none_for_unknown(ispv_sample_xlsx: Path) -> None:
    """best_code_for_family() must return None when no entry matches the given family."""
    loader = _make_loader(ispv_sample_xlsx)
    # "education" family has no entry in the fixture dataset
    result = loader.best_code_for_family("education")

    assert result is None


# ---------------------------------------------------------------------------
# _resolve() family fallback — Bug 1 whitelist
# ---------------------------------------------------------------------------


def test_family_fallback_uses_highest_pocet(tmp_path: Path) -> None:
    """Family fallback in _resolve() must return the row with max pocet, not first-match.

    Build a dataset with two 'other' family codes (major group 1 → 'other'):
      1120 (pocet=50) and 1111 (pocet=200).
    Lookup code '4111' → family 'other' (major group 4, not in _ISCO_FAMILY).
    No 4/3/2/1-digit prefix match against 1xxx entries → pure family fallback.
    Must select 1111 (highest pocet), not 1120 (first-match).
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "MZS-M8r"
    for _ in range(8):
        ws.append(["HEADER", None, None, None, None, None, None, None])
    # Lower-pocet entry first — dict ordering must not be trusted.
    ws.append(["1120 Manažeři výroby", 0.050, 85000, 55000, 68000, 110000, 145000, 92000])
    ws.append(["1111 Zákonodárci", 0.200, 95000, 60000, 75000, 120000, 160000, 100000])
    path = tmp_path / "family_fallback_test.xlsx"
    wb.save(path)

    loader = _make_loader(path)

    # 4111 → family "other" (major group 4 not in _ISCO_FAMILY → fallback "other")
    # No 4/3/2/1-digit prefix match against 1111 or 1120 → must use family fallback
    result = loader.lookup("4111")

    assert result.isco_match_level == "family_fallback"
    # Must pick 1111 (pocet=200), not 1120 (pocet=50)
    assert result.isco_code == "1111"
